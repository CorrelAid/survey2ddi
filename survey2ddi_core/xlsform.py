"""Parse XLSForm + extract a normalized variable list (source-agnostic)."""

from pathlib import Path

import openpyxl

from survey2ddi_core.types import Choice, Variable
from survey2ddi_core._generated.type_mappings import (
    TYPE_MAP,
    MEASURE_MAP,
    METADATA_TYPES,
    NON_DDI_EMITTABLE_TYPES,
    STRUCTURAL_TYPES,
)

# NON_DDI_EMITTABLE_TYPES = types lacking a ddi.intrvl block (note, geo*,
# media, phonenumber, csv-external, ...). LS_UNSUPPORTED_TYPES is *not* used
# here: it lists types unsupported by the LimeSurvey TSV parser, many of
# which are still DDI-emittable (range, acknowledge, select_*_from_file).
# `note` is excluded from skips so survey2ddi_core.notes.classify_notes can
# route note rows to <preQTxt> on the next data var, or to <notes> under
# <stdyDscr>. The CSV emitter drops notes separately.
SKIP_TYPES: set[str] = (METADATA_TYPES | STRUCTURAL_TYPES | NON_DDI_EMITTABLE_TYPES) - {"note"}


def parse_xlsform(
    path: Path,
) -> tuple[list[dict], dict[str, tuple[Choice, ...]], dict]:
    """Parse an XLSForm xlsx into (survey_rows, choices_by_list, settings).

    Handles any language suffix on label columns (e.g. ``label::Deutsch``).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def _read_sheet(name: str) -> list[dict]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = list(rows[0])
        return [dict(zip(headers, row)) for row in rows[1:]]

    survey_rows = _read_sheet("survey")
    choices_raw = _read_sheet("choices")
    settings_rows = _read_sheet("settings")
    wb.close()

    # Build choices lookup: list_name → tuple[Choice, ...]
    accum: dict[str, list[Choice]] = {}
    choice_label_col = _find_label_col([r.keys() for r in choices_raw[:1]])
    for row in choices_raw:
        ln = row.get("list_name")
        if not ln:
            continue
        accum.setdefault(str(ln), []).append(
            Choice(
                name=str(row.get("name", "")),
                label=str(row.get(choice_label_col, "") or ""),
            )
        )
    choices_by_list: dict[str, tuple[Choice, ...]] = {
        k: tuple(v) for k, v in accum.items()
    }

    settings = settings_rows[0] if settings_rows else {}

    return survey_rows, choices_by_list, settings


def resolve_title(override: str | None, settings: dict, fallback: str = "") -> str:
    """Resolve survey title from override, parsed settings.form_title, or fallback.

    Returns the first non-empty value. Used by CLIs to skip remote API lookups
    when the XLSForm/schema already carries a title.
    """
    return (
        (override or "").strip()
        or str(settings.get("form_title") or "").strip()
        or str(fallback or "").strip()
    )


def _find_label_col(header_sets: list) -> str:
    """Return the best label column name (prefer ``label::*``, fall back to ``label``).

    Limitation: for multi-language forms only the *first* ``label::*`` column is
    used.  There is currently no way to select a specific language.  If this
    matters, ensure the desired language column appears first in the XLSForm.
    """
    if not header_sets:
        return "label"
    headers = list(header_sets[0])
    for h in headers:
        if h and str(h).startswith("label::"):
            return str(h)
    return "label"


def extract_variables(
    survey_rows: list[dict],
    choices_by_list: dict[str, tuple[Choice, ...]] | dict[str, list[dict]],
) -> list[Variable]:
    """Extract a flat list of ``Variable`` from parsed XLSForm data.

    Source-agnostic — usable by both xlsx and DDI XML builders.
    ``choices_by_list`` values may be ``tuple[Choice, ...]`` (canonical) or
    raw ``list[dict]`` with ``name``/``label`` keys (for convenience in
    callers building schemas inline).
    """
    choices_by_list = {
        ln: tuple(
            c if isinstance(c, Choice)
            else Choice(name=str(c.get("name", "")), label=str(c.get("label", "") or ""))
            for c in items
        )
        for ln, items in choices_by_list.items()
    }

    label_col = _find_label_col(
        [r.keys() for r in survey_rows[:1]] if survey_rows else []
    )

    variables: list[Variable] = []
    group_stack: list[str] = []
    # Track group metadata: name → {label, appearance}
    group_meta: dict[str, dict] = {}

    for row in survey_rows:
        raw_type = str(row.get("type") or "").strip()
        if not raw_type:
            continue

        base_type = raw_type.split()[0]

        if base_type == "begin_group":
            gname = str(row.get("name", ""))
            group_stack.append(gname)
            group_meta[gname] = {
                "label": str(row.get(label_col, "") or ""),
                "appearance": str(row.get("appearance", "") or "").lower(),
            }
            continue
        if base_type == "end_group":
            if group_stack:
                group_stack.pop()
            continue
        if base_type in SKIP_TYPES:
            continue

        name = str(row.get("name", ""))
        if not name:
            continue

        # Determine standardized type and optional list name / vocab
        list_name = None
        vocab = ""
        if base_type in ("select_one", "select_multiple", "rank"):
            parts = raw_type.split(maxsplit=1)
            std_type = base_type
            list_name = parts[1] if len(parts) > 1 else None
        elif base_type in ("select_one_from_file", "select_multiple_from_file"):
            # Second token is the CSV filename; the vocab is the filename stem.
            parts = raw_type.split(maxsplit=1)
            std_type = base_type
            filename = parts[1] if len(parts) > 1 else ""
            if filename.lower().endswith(".csv"):
                vocab = filename[: -len(".csv")]
            else:
                vocab = filename
        else:
            std_type = TYPE_MAP.get(base_type, base_type)

        # Resolve choices
        choices = choices_by_list.get(list_name, ()) if list_name else ()
        values_str = "|".join(f"{c.name}={c.label}" for c in choices)

        group = "/".join(group_stack) if group_stack else ""
        data_key = f"{group}/{name}" if group else name

        # Infer measurement level
        measure = MEASURE_MAP.get(std_type, "")
        if std_type == "select_one":
            appearance = str(row.get("appearance", "") or "").lower()
            if "likert" in (list_name or "").lower() or "likert" in appearance:
                measure = "ordinal"

        # Resolve immediate group metadata (innermost group)
        cur_group = group_stack[-1] if group_stack else ""
        gm = group_meta.get(cur_group, {})

        variables.append(Variable(
            name=name,
            group=group,
            group_label=gm.get("label", ""),
            group_appearance=gm.get("appearance", ""),
            label=str(row.get(label_col, "") or ""),
            type=std_type,
            measure=measure,
            list_name=list_name or "",
            vocab=vocab,
            choices=choices,
            values=values_str,
            required=str(row.get("required", "false") or "false").lower(),
            source_type=raw_type,
            data_key=data_key,
        ))

    return variables


