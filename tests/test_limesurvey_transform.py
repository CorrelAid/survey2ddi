"""Tests for limesurvey2ddi.transform — response normalisation and DDI output."""

import subprocess
from datetime import date
from pathlib import Path
from xml.etree.ElementTree import fromstring

import pytest
from openpyxl import Workbook

from kobo2ddi.transform import extract_variables
from limesurvey2ddi.transform import build_ddi_xml, build_workbook, normalize_responses

NS = {"ddi": "ddi:codebook:2_5"}
SCHEMA_PATH = Path(__file__).parent / "schemas" / "codebook.xsd"

# ---------------------------------------------------------------------------
# Fixtures — a minimal LimeSurvey-style survey
# ---------------------------------------------------------------------------

LIME_SURVEY_ROWS = [
    {"type": "select_one haeufigkeit", "name": "haeufigkeit", "label": "How often?", "required": "true", "appearance": None},
    {"type": "select_multiple bereiche", "name": "bereiche", "label": "Which areas?", "required": "false", "appearance": None},
    {"type": "select_one likert5", "name": "beruf_post", "label": "Career clarity", "required": "true", "appearance": None},
    {"type": "text", "name": "am_meisten_gebracht", "label": "What helped most?", "required": "false", "appearance": None},
    {"type": "select_one nps", "name": "nps_score", "label": "NPS", "required": "true", "appearance": None},
]

LIME_CHOICES = {
    "haeufigkeit": [
        {"name": "1", "label": "Once"},
        {"name": "2", "label": "Sometimes"},
    ],
    "bereiche": [
        {"name": "holz", "label": "Wood workshop"},
        {"name": "metall", "label": "Metal workshop"},
        {"name": "textil", "label": "Textile workshop"},
        {"name": "digital", "label": "Digital workshop"},
    ],
    "likert5": [
        {"name": "1", "label": "Strongly disagree"},
        {"name": "5", "label": "Strongly agree"},
    ],
    "nps": [{"name": str(i), "label": str(i)} for i in range(11)],
}

LIME_SETTINGS = {"id_string": "lime_test", "version": "1.0", "default_language": "Deutsch"}

# A single LimeSurvey response row — note the export quirks:
#   - underscores stripped: beruf_post → berufpost, am_meisten_gebracht → ammeistengebracht, nps_score → npsscore
#   - select_multiple as sub-columns with truncated option codes (5 chars max)
#   - metadata fields mixed in
LIME_RESPONSE_ROW = {
    "id": "1",
    "submitdate": "2025-01-01",
    "startdate": "2025-01-01",
    "datestamp": "2025-01-01",
    "lastpage": "5",
    "startlanguage": "de",
    "seed": "42",
    "haeufigkeit": "Once",
    "bereiche[holz]": "Yes",
    "bereiche[metal]": "No",     # metall truncated to metal
    "bereiche[texti]": "Yes",    # textil truncated to texti
    "bereiche[digit]": "No",     # digital truncated to digit
    "berufpost": "Strongly agree",
    "ammeistengebracht": "It was great",
    "npsscore": "9",
}


@pytest.fixture
def lime_variables():
    return extract_variables(LIME_SURVEY_ROWS, LIME_CHOICES)


@pytest.fixture
def lime_responses():
    return [LIME_RESPONSE_ROW]


@pytest.fixture
def lime_form_path(tmp_path):
    """Write a minimal LimeSurvey XLSForm xlsx and return its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "survey"
    ws.append(["type", "name", "label", "required", "appearance"])
    for row in LIME_SURVEY_ROWS:
        ws.append([row["type"], row["name"], row.get("label"), row.get("required"), row.get("appearance")])

    ws_c = wb.create_sheet("choices")
    ws_c.append(["list_name", "name", "label"])
    for list_name, choices in LIME_CHOICES.items():
        for c in choices:
            ws_c.append([list_name, c["name"], c["label"]])

    ws_s = wb.create_sheet("settings")
    ws_s.append(["id_string", "version", "default_language"])
    ws_s.append([LIME_SETTINGS["id_string"], LIME_SETTINGS["version"], LIME_SETTINGS["default_language"]])

    path = tmp_path / "form.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# normalize_responses
# ---------------------------------------------------------------------------


class TestNormalizeResponses:
    def test_simple_field_lookup(self, lime_variables, lime_responses):
        result = normalize_responses(lime_variables, lime_responses)
        assert result[0]["haeufigkeit"] == "Once"

    def test_underscore_stripping(self, lime_variables, lime_responses):
        """beruf_post in form → berufpost in LimeSurvey export."""
        result = normalize_responses(lime_variables, lime_responses)
        assert result[0]["beruf_post"] == "Strongly agree"

    def test_underscore_stripping_multiple(self, lime_variables, lime_responses):
        """am_meisten_gebracht → ammeistengebracht, nps_score → npsscore."""
        result = normalize_responses(lime_variables, lime_responses)
        assert result[0]["am_meisten_gebracht"] == "It was great"
        assert result[0]["nps_score"] == "9"

    def test_select_multiple_yes_becomes_code(self, lime_variables, lime_responses):
        """bereiche[holz]="Yes" → "holz" is included in the result."""
        result = normalize_responses(lime_variables, lime_responses)
        selected = result[0]["bereiche"].split()
        assert "holz" in selected

    def test_select_multiple_no_excluded(self, lime_variables, lime_responses):
        """bereiche[metal]="No" → "metall" should not appear in the result."""
        result = normalize_responses(lime_variables, lime_responses)
        selected = result[0]["bereiche"].split()
        assert "metall" not in selected
        assert "metal" not in selected

    def test_select_multiple_truncated_code_recovered(self, lime_variables, lime_responses):
        """bereiche[texti]="Yes" → recovers full code "textil" via prefix match."""
        result = normalize_responses(lime_variables, lime_responses)
        selected = result[0]["bereiche"].split()
        assert "textil" in selected

    def test_select_multiple_all_no(self, lime_variables):
        """All sub-columns "No" → empty string."""
        row = {
            "bereiche[holz]": "No", "bereiche[metal]": "No",
            "bereiche[texti]": "No", "bereiche[digit]": "No",
        }
        result = normalize_responses(lime_variables, [row])
        assert result[0]["bereiche"] == ""

    def test_select_multiple_ambiguous_prefix_raises(self):
        """Two choice codes sharing the same 5-char prefix → ValueError, not silent wrong data."""
        from kobo2ddi.transform import extract_variables
        rows = [{"type": "select_multiple tags", "name": "tags", "label": "Tags", "required": "false"}]
        # "optie_a" and "optie_b" both truncate to "optie" in LimeSurvey bracket keys
        choices = {"tags": [{"name": "optie_a", "label": "A"}, {"name": "optie_b", "label": "B"}]}
        variables = extract_variables(rows, choices)
        response = [{"tags[optie]": "Yes"}]
        with pytest.raises(ValueError, match="Ambiguous"):
            normalize_responses(variables, response)

    def test_select_multiple_unknown_bracket_key_warns(self):
        """Bracket key with no matching choice code → warning, raw key used."""
        from kobo2ddi.transform import extract_variables
        rows = [{"type": "select_multiple opts", "name": "opts", "label": "Opts", "required": "false"}]
        choices = {"opts": [{"name": "alpha", "label": "Alpha"}]}
        variables = extract_variables(rows, choices)
        response = [{"opts[zzzzz]": "Yes"}]  # "zzzzz" matches nothing
        with pytest.warns(UserWarning, match="did not match any XLSForm choice code"):
            result = normalize_responses(variables, response)
        assert result[0]["opts"] == "zzzzz"  # raw key preserved

    def test_select_multiple_multiple_selected(self, lime_variables):
        """Multiple "Yes" sub-columns → space-separated codes."""
        row = {
            "bereiche[holz]": "Yes", "bereiche[metal]": "Yes",
            "bereiche[texti]": "No", "bereiche[digit]": "No",
        }
        result = normalize_responses(lime_variables, [row])
        selected = result[0]["bereiche"].split()
        assert "holz" in selected
        assert "metall" in selected
        assert len(selected) == 2

    def test_metadata_fields_not_in_output(self, lime_variables, lime_responses):
        """LimeSurvey metadata fields (id, submitdate, etc.) are not present."""
        result = normalize_responses(lime_variables, lime_responses)
        for meta in ("id", "submitdate", "startdate", "datestamp", "seed"):
            assert meta not in result[0]

    def test_keyed_by_data_key(self, lime_variables, lime_responses):
        """Output is keyed by _data_key, which equals name when no group."""
        result = normalize_responses(lime_variables, lime_responses)
        by_name = {v["name"]: v["_data_key"] for v in lime_variables}
        for var_name, data_key in by_name.items():
            assert data_key in result[0]

    def test_missing_variable_in_response_gives_empty_string(self, lime_variables):
        """Variable not present in the response row → empty string, not KeyError."""
        result = normalize_responses(lime_variables, [{}])
        assert result[0]["haeufigkeit"] == ""
        assert result[0]["bereiche"] == ""

    def test_multiple_rows(self, lime_variables):
        rows = [
            {"haeufigkeit": "Once", "bereiche[holz]": "Yes", "bereiche[metal]": "No",
             "bereiche[texti]": "No", "bereiche[digit]": "No", "berufpost": "1",
             "ammeistengebracht": "great", "npsscore": "8"},
            {"haeufigkeit": "Sometimes", "bereiche[holz]": "No", "bereiche[metal]": "Yes",
             "bereiche[texti]": "No", "bereiche[digit]": "No", "berufpost": "5",
             "ammeistengebracht": "also great", "npsscore": "10"},
        ]
        result = normalize_responses(lime_variables, rows)
        assert len(result) == 2
        assert result[0]["haeufigkeit"] == "Once"
        assert result[1]["haeufigkeit"] == "Sometimes"
        assert "holz" in result[0]["bereiche"].split()
        assert "metall" in result[1]["bereiche"].split()

    def test_empty_responses(self, lime_variables):
        assert normalize_responses(lime_variables, []) == []


# ---------------------------------------------------------------------------
# build_workbook
# ---------------------------------------------------------------------------


class TestLimeBuildWorkbook:
    def test_has_three_sheets(self, lime_form_path, lime_responses):
        wb = build_workbook("Test Survey", lime_form_path, lime_responses)
        assert wb.sheetnames == ["variables", "data", "survey_info"]

    def test_source_is_limesurvey(self, lime_form_path, lime_responses):
        wb = build_workbook("Test Survey", lime_form_path, lime_responses)
        ws = wb["survey_info"]
        info = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(2, ws.max_row + 1)}
        assert info["source"] == "limesurvey"

    def test_title_in_survey_info(self, lime_form_path, lime_responses):
        wb = build_workbook("My LimeSurvey", lime_form_path, lime_responses)
        ws = wb["survey_info"]
        info = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(2, ws.max_row + 1)}
        assert info["title"] == "My LimeSurvey"

    def test_export_date_is_today(self, lime_form_path, lime_responses):
        wb = build_workbook("T", lime_form_path, lime_responses)
        ws = wb["survey_info"]
        info = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(2, ws.max_row + 1)}
        assert info["export_date"] == date.today().isoformat()

    def test_data_sheet_has_correct_columns(self, lime_form_path, lime_responses):
        wb = build_workbook("T", lime_form_path, lime_responses)
        ws_var = wb["variables"]
        ws_data = wb["data"]
        var_names = [ws_var.cell(row=r, column=1).value for r in range(2, ws_var.max_row + 1)]
        data_header = [cell.value for cell in ws_data[1]]
        assert data_header == var_names

    def test_simple_field_in_data(self, lime_form_path, lime_responses):
        wb = build_workbook("T", lime_form_path, lime_responses)
        ws = wb["data"]
        header = [cell.value for cell in ws[1]]
        row = {header[i]: ws.cell(row=2, column=i + 1).value for i in range(len(header))}
        assert row["haeufigkeit"] == "Once"
        assert row["am_meisten_gebracht"] == "It was great"
        assert row["nps_score"] == "9"

    def test_underscore_variable_in_data(self, lime_form_path, lime_responses):
        """beruf_post (underscore in name) is correctly mapped from berufpost response."""
        wb = build_workbook("T", lime_form_path, lime_responses)
        ws = wb["data"]
        header = [cell.value for cell in ws[1]]
        row = {header[i]: ws.cell(row=2, column=i + 1).value for i in range(len(header))}
        assert row["beruf_post"] == "Strongly agree"

    def test_select_multiple_in_data(self, lime_form_path, lime_responses):
        """bereiche select_multiple shows space-separated selected codes."""
        wb = build_workbook("T", lime_form_path, lime_responses)
        ws = wb["data"]
        header = [cell.value for cell in ws[1]]
        row = {header[i]: ws.cell(row=2, column=i + 1).value for i in range(len(header))}
        selected = row["bereiche"].split()
        assert "holz" in selected
        assert "textil" in selected
        assert "metall" not in selected

    def test_empty_responses(self, lime_form_path):
        wb = build_workbook("T", lime_form_path, [])
        ws = wb["data"]
        assert ws.max_row == 1  # header only


# ---------------------------------------------------------------------------
# build_ddi_xml
# ---------------------------------------------------------------------------


class TestLimeBuildDdiXml:
    def test_returns_valid_xml(self, lime_form_path, lime_responses):
        xml = build_ddi_xml("Test Survey", lime_form_path, lime_responses)
        root = fromstring(xml)
        assert root.tag == "{ddi:codebook:2_5}codeBook"

    def test_study_title(self, lime_form_path, lime_responses):
        xml = build_ddi_xml("My LimeSurvey", lime_form_path, lime_responses)
        root = fromstring(xml)
        titl = root.find(".//ddi:stdyDscr/ddi:citation/ddi:titlStmt/ddi:titl", NS)
        assert titl.text == "My LimeSurvey"

    def test_variable_count(self, lime_form_path, lime_responses):
        xml = build_ddi_xml("T", lime_form_path, lime_responses)
        root = fromstring(xml)
        variables = root.findall(".//ddi:dataDscr/ddi:var", NS)
        assert len(variables) == len(LIME_SURVEY_ROWS)

    def test_select_multiple_has_categories(self, lime_form_path, lime_responses):
        xml = build_ddi_xml("T", lime_form_path, lime_responses)
        root = fromstring(xml)
        variables = root.findall(".//ddi:dataDscr/ddi:var", NS)
        by_name = {v.get("name"): v for v in variables}
        catgries = by_name["bereiche"].findall("ddi:catgry", NS)
        assert len(catgries) == 4
        vals = [c.find("ddi:catValu", NS).text for c in catgries]
        assert "holz" in vals
        assert "metall" in vals

    @pytest.mark.skipif(
        subprocess.run(["which", "xmllint"], capture_output=True).returncode != 0,
        reason="xmllint not available",
    )
    def test_validates_against_ddi_xsd(self, lime_form_path, lime_responses, tmp_path):
        xml = build_ddi_xml("Test Survey", lime_form_path, lime_responses)
        path = tmp_path / "lime_test.xml"
        path.write_text(xml, encoding="utf-8")
        result = subprocess.run(
            ["xmllint", "--noout", "--schema", str(SCHEMA_PATH), str(path)],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, f"XSD validation failed:\n{result.stderr}"
