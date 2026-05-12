"""Tests for limesurvey2ddi.cli — CLI argument parsing and command dispatch."""

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from limesurvey2ddi.cli import main


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------

def _make_xlsform(path: Path) -> None:
    """Write a minimal valid XLSForm xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "survey"
    ws.append(["type", "name", "label", "required"])
    ws.append(["text", "q1", "Question 1", "false"])
    wb.create_sheet("choices")
    ws_s = wb.create_sheet("settings")
    ws_s.append(["id_string", "version", "default_language"])
    ws_s.append(["cli_test", "1.0", "English"])
    wb.save(path)


@pytest.fixture
def mock_client(tmp_path):
    """Mocked LimeSurveyClient with pre-populated output directory."""
    client = MagicMock()
    client.list_surveys.return_value = [
        {"sid": "99", "surveyls_title": "Test Survey", "active": "Y"},
        {"sid": "100", "surveyls_title": "Inactive Survey", "active": "N"},
    ]
    client.validate.return_value = True
    client.pull.return_value = tmp_path / "99"

    # Pre-populate output dir so transform can run without an API call
    survey_dir = tmp_path / "99"
    survey_dir.mkdir(parents=True, exist_ok=True)
    _make_xlsform(survey_dir / "form.xlsx")
    (survey_dir / "responses.json").write_text(
        json.dumps([{"q1": "hello"}]), encoding="utf-8"
    )

    return client, tmp_path


# ---------------------------------------------------------------------------
# list
# ---------------------------------------------------------------------------

class TestCliList:
    def test_list_prints_surveys(self, mock_client, capsys):
        client, _ = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "list"])
        out = capsys.readouterr().out
        assert "99" in out
        assert "Test Survey" in out
        assert "active" in out

    def test_list_shows_inactive(self, mock_client, capsys):
        client, _ = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "list"])
        out = capsys.readouterr().out
        assert "inactive" in out

    def test_list_empty(self, capsys):
        client = MagicMock()
        client.list_surveys.return_value = []
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "list"])
        assert "No surveys found" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# pull
# ---------------------------------------------------------------------------

class TestCliPull:
    def test_pull_calls_client(self, mock_client):
        client, tmp_path = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "pull", "99", "-o", str(tmp_path)])
        client.pull.assert_called_once_with(99, output_dir=tmp_path)

    def test_pull_uses_default_output_when_omitted(self, mock_client):
        client, _ = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "pull", "99"])
        client.pull.assert_called_once_with(99, output_dir=None)


# ---------------------------------------------------------------------------
# validate
# ---------------------------------------------------------------------------

class TestCliValidate:
    def test_validate_calls_client(self, mock_client):
        client, tmp_path = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "validate", "99", "-o", str(tmp_path)])
        client.validate.assert_called_once_with(99, output_dir=tmp_path)

    def test_validate_exits_on_failure(self, mock_client):
        client, tmp_path = mock_client
        client.validate.return_value = False
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            with pytest.raises(SystemExit) as exc:
                main(["--username", "u", "--password", "p", "validate", "99", "-o", str(tmp_path)])
        assert exc.value.code == 1


# ---------------------------------------------------------------------------
# transform
# ---------------------------------------------------------------------------

class TestCliTransform:
    def test_transform_creates_csv_and_xml(self, mock_client):
        client, tmp_path = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "transform", "99", "-o", str(tmp_path)])
        survey_dir = tmp_path / "99"
        assert (survey_dir / "99.csv").exists()
        assert (survey_dir / "99.xml").exists()

    def test_transform_uses_title_arg(self, mock_client, capsys):
        client, tmp_path = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "transform", "99",
                  "--title", "My Survey", "-o", str(tmp_path)])
        assert "Wrote" in capsys.readouterr().out
        xml = (tmp_path / "99" / "99.xml").read_text()
        assert "<titl>My Survey</titl>" in xml

    def test_transform_defaults_title_to_survey_id(self, mock_client, tmp_path):
        client, tmp_path = mock_client
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "transform", "99", "-o", str(tmp_path)])
        xml = (tmp_path / "99" / "99.xml").read_text()
        assert "<titl>99</titl>" in xml

    def test_transform_with_tsv_schema(self, mock_client, tmp_path):
        client, _ = mock_client
        survey_dir = tmp_path / "101"
        survey_dir.mkdir(parents=True)
        # Use survey.tsv instead of form.xlsx
        tsv_path = survey_dir / "survey.tsv"
        tsv_path.write_text(
            "class\ttype/scale\tname\ttext\tlanguage\n"
            "S\t\t\tTest Survey\ten\n"
            "G\t\tG1\tGroup 1\ten\n"
            "Q\tT\tq1\tQuestion 1\ten\n",
            encoding="utf-8",
        )
        (survey_dir / "responses.json").write_text(
            json.dumps([{"q1": "hello"}]), encoding="utf-8"
        )

        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main(["--username", "u", "--password", "p", "transform", "101", "-o", str(tmp_path)])

        assert (survey_dir / "101.csv").exists()
        assert (survey_dir / "101.xml").exists()
        xml = (survey_dir / "101.xml").read_text()
        assert "q1" in xml

    def test_transform_with_explicit_schema_arg(self, mock_client, tmp_path):
        client, _ = mock_client
        survey_dir = tmp_path / "102"
        survey_dir.mkdir(parents=True)
        custom_schema = tmp_path / "custom.tsv"
        custom_schema.write_text(
            "class\ttype/scale\tname\ttext\tlanguage\n"
            "S\t\t\tTest Survey\ten\n"
            "G\t\tG1\tGroup 1\ten\n"
            "Q\tT\tq2\tQuestion 2\ten\n",
            encoding="utf-8",
        )
        (survey_dir / "responses.json").write_text(
            json.dumps([{"q2": "world"}]), encoding="utf-8"
        )

        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main([
                "--username", "u", "--password", "p", "transform", "102",
                "--schema", str(custom_schema),
                "-o", str(tmp_path)
            ])

        assert (survey_dir / "102.xml").exists()
        xml = (survey_dir / "102.xml").read_text()
        assert "q2" in xml

    def test_transform_without_client_instantiation(self, mock_client, tmp_path):
        """Offline transform: no client constructed if files are present."""
        client, _ = mock_client
        survey_dir = tmp_path / "103"
        survey_dir.mkdir(parents=True)
        _make_xlsform(survey_dir / "form.xlsx")
        (survey_dir / "responses.json").write_text("[]", encoding="utf-8")

        # Mock LimeSurveyClient to fail if instantiated
        lime_client_mock = MagicMock(side_effect=AssertionError("LimeSurveyClient must not be instantiated"))
        with patch("limesurvey2ddi.cli.LimeSurveyClient", lime_client_mock):
            main(["transform", "103", "-o", str(tmp_path)])

        lime_client_mock.assert_not_called()
        assert (survey_dir / "103.xml").exists()

    def test_transform_with_csv_data(self, mock_client, tmp_path):
        """--data uses a CSV file instead of responses.json."""
        client, _ = mock_client
        survey_dir = tmp_path / "104"
        survey_dir.mkdir(parents=True)
        _make_xlsform(survey_dir / "form.xlsx")
        
        # LimeSurvey CSV with question codes as headers and semicolon
        csv_path = tmp_path / "responses.csv"
        csv_path.write_text("id;submitdate;q1\n1;2025-06-01;Alice\n2;2025-06-02;Bob", encoding="utf-8")

        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            main([
                "transform", "104",
                "-o", str(tmp_path),
                "--data", str(csv_path),
                "--title", "CSV Survey",
            ])
        
        assert (survey_dir / "104.csv").exists()
        assert (survey_dir / "104.xml").exists()
        csv_text = (survey_dir / "104.csv").read_text()
        assert "Alice" in csv_text
        assert "Bob" in csv_text

    def test_transform_missing_form_exits(self, mock_client, tmp_path):
        client, _ = mock_client
        empty_dir = tmp_path / "empty"
        empty_dir.mkdir()
        # No form.xlsx present
        (empty_dir / "77").mkdir()
        (empty_dir / "77" / "responses.json").write_text("[]")
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            with pytest.raises(SystemExit) as exc:
                main(["--username", "u", "--password", "p", "transform", "77",
                      "-o", str(empty_dir)])
        assert exc.value.code == 1

    def test_transform_missing_responses_exits(self, mock_client, tmp_path):
        client, _ = mock_client
        survey_dir = tmp_path / "88"
        survey_dir.mkdir()
        _make_xlsform(survey_dir / "form.xlsx")
        # No responses.json present
        with patch("limesurvey2ddi.cli.LimeSurveyClient", return_value=client):
            with pytest.raises(SystemExit) as exc:
                main(["--username", "u", "--password", "p", "transform", "88",
                      "-o", str(tmp_path)])
        assert exc.value.code == 1


# ---------------------------------------------------------------------------
# no command
# ---------------------------------------------------------------------------

class TestCliNoCommand:
    def test_no_command_exits(self):
        with pytest.raises(SystemExit):
            main(["--username", "u", "--password", "p"])
