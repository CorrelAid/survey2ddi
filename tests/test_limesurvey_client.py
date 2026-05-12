"""Tests for limesurvey2ddi.client — LimeSurveyClient with mocked HTTP."""

import base64
import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import httpx
import pytest
from openpyxl import Workbook

from limesurvey2ddi.client import LimeSurveyClient


# -- Helpers -----------------------------------------------------------------


def _mock_response(json_data=None, status_code=200):
    resp = MagicMock(spec=httpx.Response)
    resp.status_code = status_code
    resp.json.return_value = json_data
    resp.raise_for_status = MagicMock()
    return resp


def _rpc_response(result, error=None):
    return _mock_response(json_data={"id": 1, "result": result, "error": error})


def _b64_responses(rows: list[dict]) -> str:
    """Encode a list of response dicts as the base64 JSON LimeSurvey returns."""
    payload = {"responses": rows}
    return base64.b64encode(json.dumps(payload).encode()).decode()


@pytest.fixture
def client():
    """LimeSurveyClient with fake credentials; session key pre-set to skip auth call."""
    c = LimeSurveyClient(
        server_url="https://lime.example.org",
        username="testuser",
        password="testpass",
    )
    c._session_key = "fake-session-key"
    return c


def _make_xlsform(path: Path, variables: list[tuple]) -> None:
    """Write a minimal XLSForm xlsx. variables is [(type, name), ...]."""
    wb = Workbook()
    ws = wb.active
    ws.title = "survey"
    ws.append(["type", "name", "label"])
    for vtype, vname in variables:
        ws.append([vtype, vname, vname])
    wb.create_sheet("choices")
    wb.create_sheet("settings")
    wb.save(path)


# -- Init --------------------------------------------------------------------


class TestLimeSurveyClientInit:
    def test_requires_credentials(self):
        with patch.dict("os.environ", {}, clear=True):
            with pytest.raises(ValueError, match="Username and password required"):
                LimeSurveyClient(server_url="https://x.org")

    def test_credentials_from_arguments(self):
        c = LimeSurveyClient(
            server_url="https://x.org", username="u", password="p"
        )
        assert c._username == "u"
        assert c._password == "p"

    def test_credentials_from_env(self):
        with patch.dict("os.environ", {"LIME_USERNAME": "eu", "LIME_PASSWORD": "ep"}):
            c = LimeSurveyClient(server_url="https://x.org")
            assert c._username == "eu"

    def test_server_url_trailing_slash_stripped(self):
        c = LimeSurveyClient(server_url="https://x.org/", username="u", password="p")
        assert c.server_url == "https://x.org"

    def test_default_server_url(self):
        with patch.dict("os.environ", {"LIME_USERNAME": "u", "LIME_PASSWORD": "p",
                                        "LIME_SERVER_URL": ""}):
            c = LimeSurveyClient()
            assert "lime.correlaid.org" in c.server_url


# -- _rpc transport ----------------------------------------------------------


class TestRpc:
    def test_session_key_prepended_automatically(self, client):
        """Session key should be prepended to params for non-auth methods."""
        mock_post = MagicMock(return_value=_rpc_response("ok"))
        with patch.object(client._http, "post", mock_post):
            client._rpc("list_surveys", "user")
        payload = json.loads(mock_post.call_args.kwargs["content"])
        assert payload["params"][0] == "fake-session-key"
        assert payload["params"][1] == "user"

    def test_session_key_not_prepended_for_auth_methods(self, client):
        """get_session_key and release_session_key receive no auto-injected key."""
        mock_post = MagicMock(return_value=_rpc_response("new-key"))
        with patch.object(client._http, "post", mock_post):
            client._rpc("get_session_key", "user", "pass")
        payload = json.loads(mock_post.call_args.kwargs["content"])
        assert payload["params"] == ["user", "pass"]

    def test_lazy_session_key_acquisition(self):
        """Session key is obtained lazily on first non-auth RPC call."""
        c = LimeSurveyClient(server_url="https://x.org", username="u", password="p")
        assert c._session_key is None
        session_resp = _rpc_response("acquired-key")
        method_resp = _rpc_response([{"sid": "1", "surveyls_title": "S"}])
        with patch.object(c._http, "post", side_effect=[session_resp, method_resp]):
            c.list_surveys()
        assert c._session_key == "acquired-key"

    def test_raises_on_api_error(self, client):
        mock_post = MagicMock(return_value=_rpc_response(None, error="Invalid key"))
        with patch.object(client._http, "post", mock_post):
            with pytest.raises(RuntimeError, match="Invalid key"):
                client._rpc("list_surveys", "u")


# -- list_surveys ------------------------------------------------------------


class TestListSurveys:
    def test_returns_surveys(self, client):
        surveys = [
            {"sid": "1", "surveyls_title": "Survey A", "active": "Y"},
            {"sid": "2", "surveyls_title": "Survey B", "active": "N"},
        ]
        with patch.object(client._http, "post", return_value=_rpc_response(surveys)):
            result = client.list_surveys()
        assert len(result) == 2
        assert result[0]["surveyls_title"] == "Survey A"

    def test_returns_empty_list_when_no_surveys(self, client):
        with patch.object(
            client._http, "post",
            return_value=_rpc_response({"status": "No surveys found"}),
        ):
            assert client.list_surveys() == []


# -- get_responses -----------------------------------------------------------


class TestGetResponses:
    def test_flat_format(self, client):
        """Current LimeSurvey server returns flat dicts in responses list."""
        rows = [
            {"id": "1", "submitdate": "2025-01-01", "q1": "answer_a"},
            {"id": "2", "submitdate": "2025-01-02", "q1": "answer_b"},
        ]
        with patch.object(
            client._http, "post",
            return_value=_rpc_response(_b64_responses(rows)),
        ):
            result = client.get_responses(12345)
        assert len(result) == 2
        assert result[0]["q1"] == "answer_a"
        assert result[1]["q1"] == "answer_b"

    def test_nested_format(self, client):
        """Some LimeSurvey versions wrap each row as {'n': {field: value}}."""
        rows = [
            {"1": {"id": "1", "submitdate": "2025-01-01", "q1": "answer_a"}},
            {"2": {"id": "2", "submitdate": "2025-01-02", "q1": "answer_b"}},
        ]
        with patch.object(
            client._http, "post",
            return_value=_rpc_response(_b64_responses(rows)),
        ):
            result = client.get_responses(12345)
        assert len(result) == 2
        assert result[0]["q1"] == "answer_a"

    def test_no_responses(self, client):
        with patch.object(
            client._http, "post",
            return_value=_rpc_response({"status": "No Data"}),
        ):
            assert client.get_responses(12345) == []

    def test_non_dict_items_skipped(self, client):
        """Non-dict entries in responses array are silently skipped (line 116)."""
        rows = [
            {"id": "1", "q1": "a"},
            "unexpected_string",   # non-dict — should be ignored
            {"id": "2", "q1": "b"},
        ]
        with patch.object(
            client._http, "post",
            return_value=_rpc_response(_b64_responses(rows)),
        ):
            result = client.get_responses(12345)
        assert len(result) == 2
        assert result[0]["q1"] == "a"
        assert result[1]["q1"] == "b"

    def test_response_keys_preserved(self, client):
        """Response dict keys (including metadata) are preserved as-is."""
        rows = [{"id": "1", "submitdate": "2025-01-01", "haeufigkeit": "often"}]
        with patch.object(
            client._http, "post",
            return_value=_rpc_response(_b64_responses(rows)),
        ):
            result = client.get_responses(99)
        assert "haeufigkeit" in result[0]
        assert "submitdate" in result[0]


# -- pull --------------------------------------------------------------------


class TestPull:
    def test_saves_responses_json(self, client, tmp_path):
        rows = [{"id": "1", "q1": "a"}]
        with patch.object(
            client._http, "post",
            return_value=_rpc_response(_b64_responses(rows)),
        ):
            out = client.pull(99, output_dir=tmp_path)
        saved = json.loads((out / "responses.json").read_text())
        assert len(saved) == 1
        assert saved[0]["q1"] == "a"

    def test_prints_reminder_when_no_tsv(self, client, tmp_path, capsys):
        rows = [{"id": "1", "q1": "a"}]
        with patch.object(
            client._http, "post",
            return_value=_rpc_response(_b64_responses(rows)),
        ):
            client.pull(99, output_dir=tmp_path)
        out = capsys.readouterr().out
        assert "survey.tsv" in out
