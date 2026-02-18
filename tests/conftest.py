"""Shared test fixtures."""

import pytest
from openpyxl import Workbook
from pathlib import Path


# -- Reusable survey data as plain dicts (same shape as parse_xlsform output) --

SURVEY_ROWS = [
    {"type": "begin_group", "name": "demo", "label": None, "required": "false", "appearance": None},
    {"type": "text", "name": "full_name", "label": "Full name", "required": "true", "appearance": None},
    {"type": "integer", "name": "age", "label": "Age", "required": "true", "appearance": None},
    {"type": "select_one gender", "name": "gender", "label": "Gender", "required": "false", "appearance": None},
    {"type": "select_multiple hobbies", "name": "hobbies", "label": "Hobbies", "required": "false", "appearance": None},
    {"type": "end_group", "name": None, "label": None, "required": None, "appearance": None},
    {"type": "begin_group", "name": "feedback", "label": None, "required": "false", "appearance": None},
    {"type": "select_one likert", "name": "satisfaction", "label": "How satisfied are you?", "required": "true", "appearance": "likert"},
    {"type": "range", "name": "score", "label": "Score (0-100)", "required": "false", "appearance": None},
    {"type": "decimal", "name": "rating", "label": "Rating", "required": "false", "appearance": None},
    {"type": "date", "name": "visit_date", "label": "Date of visit", "required": "false", "appearance": None},
    {"type": "calculate", "name": "calc_field", "label": "Computed", "required": "false", "appearance": None},
    {"type": "end_group", "name": None, "label": None, "required": None, "appearance": None},
    {"type": "note", "name": "thanks", "label": "Thank you!", "required": "false", "appearance": None},
]

CHOICES_BY_LIST = {
    "gender": [
        {"name": "m", "label": "Male"},
        {"name": "f", "label": "Female"},
        {"name": "other", "label": "Other"},
    ],
    "hobbies": [
        {"name": "sports", "label": "Sports"},
        {"name": "music", "label": "Music"},
        {"name": "reading", "label": "Reading"},
    ],
    "likert": [
        {"name": "1", "label": "Very unsatisfied"},
        {"name": "2", "label": "Unsatisfied"},
        {"name": "3", "label": "Neutral"},
        {"name": "4", "label": "Satisfied"},
        {"name": "5", "label": "Very satisfied"},
    ],
}

SETTINGS = {
    "id_string": "test_survey_2025",
    "version": "1.0",
    "default_language": "English",
}

SUBMISSIONS = [
    {
        "demo/full_name": "Alice",
        "demo/age": "30",
        "demo/gender": "f",
        "demo/hobbies": "sports music",
        "feedback/satisfaction": "4",
        "feedback/score": "75",
        "feedback/rating": "4.5",
        "feedback/visit_date": "2025-06-01",
        "feedback/calc_field": "42",
    },
    {
        "demo/full_name": "Bob",
        "demo/age": "25",
        "demo/gender": "m",
        "demo/hobbies": "reading",
        "feedback/satisfaction": "5",
        "feedback/score": "90",
        "feedback/rating": "4.8",
        "feedback/visit_date": "2025-06-02",
        "feedback/calc_field": "99",
    },
]


@pytest.fixture
def survey_rows():
    return SURVEY_ROWS


@pytest.fixture
def choices_by_list():
    return CHOICES_BY_LIST


@pytest.fixture
def settings():
    return SETTINGS


@pytest.fixture
def submissions():
    return SUBMISSIONS


@pytest.fixture
def xlsform_path(tmp_path):
    """Create a minimal XLSForm xlsx file and return its path."""
    wb = Workbook()

    # survey sheet
    ws = wb.active
    ws.title = "survey"
    ws.append(["type", "name", "label::English", "required", "appearance"])
    for row in SURVEY_ROWS:
        ws.append([row["type"], row["name"], row.get("label"), row.get("required"), row.get("appearance")])

    # choices sheet
    ws_c = wb.create_sheet("choices")
    ws_c.append(["list_name", "name", "label::English"])
    for list_name, choices in CHOICES_BY_LIST.items():
        for c in choices:
            ws_c.append([list_name, c["name"], c["label"]])

    # settings sheet
    ws_s = wb.create_sheet("settings")
    ws_s.append(["id_string", "version", "default_language"])
    ws_s.append([SETTINGS["id_string"], SETTINGS["version"], SETTINGS["default_language"]])

    path = tmp_path / "form.xlsx"
    wb.save(path)
    return path
