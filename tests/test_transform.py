"""Tests for kobo2ddi.transform — variable extraction, XLSForm parsing, xlsx output."""

from datetime import date

from kobo2ddi.transform import extract_variables, parse_xlsform, build_workbook


# -- extract_variables -------------------------------------------------------


class TestExtractVariables:
    def test_returns_only_data_carrying_types(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        names = [v["name"] for v in variables]
        # data-carrying fields present
        assert "full_name" in names
        assert "age" in names
        assert "gender" in names
        assert "satisfaction" in names
        # non-data types excluded
        assert "thanks" not in names  # note
        assert "demo" not in names  # begin_group name

    def test_group_tracking(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["group"] == "demo"
        assert by_name["satisfaction"]["group"] == "feedback"

    def test_data_key_includes_group(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["_data_key"] == "demo/full_name"
        assert by_name["satisfaction"]["_data_key"] == "feedback/satisfaction"

    def test_type_mapping(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["type"] == "string"
        assert by_name["age"]["type"] == "integer"
        assert by_name["gender"]["type"] == "select_one"
        assert by_name["hobbies"]["type"] == "select_multiple"
        assert by_name["satisfaction"]["type"] == "select_one"
        assert by_name["score"]["type"] == "range"
        assert by_name["rating"]["type"] == "decimal"
        assert by_name["visit_date"]["type"] == "date"
        assert by_name["calc_field"]["type"] == "calculate"

    def test_choices_resolved(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        gender = by_name["gender"]
        assert len(gender["choices"]) == 3
        assert gender["values"] == "m=Male|f=Female|other=Other"

    def test_no_choices_for_non_select(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["choices"] == []
        assert by_name["full_name"]["values"] == ""

    def test_source_type_preserved(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["gender"]["source_type"] == "select_one gender"
        assert by_name["full_name"]["source_type"] == "text"

    def test_unknown_type_passed_through(self):
        rows = [{"type": "newtype", "name": "x", "label": "X", "required": "false"}]
        variables = extract_variables(rows, {})
        assert variables[0]["type"] == "newtype"

    def test_empty_survey_rows(self):
        assert extract_variables([], {}) == []

    def test_skipped_types_not_included(self):
        rows = [
            {"type": "note", "name": "n1", "label": "Note"},
            {"type": "geopoint", "name": "loc", "label": "Location"},
            {"type": "image", "name": "photo", "label": "Photo"},
            {"type": "start", "name": "start", "label": "Start"},
        ]
        assert extract_variables(rows, {}) == []


# -- parse_xlsform -----------------------------------------------------------


class TestParseXlsform:
    def test_parses_survey_sheet(self, xlsform_path):
        survey_rows, _, _ = parse_xlsform(xlsform_path)
        types = [r["type"] for r in survey_rows]
        assert "text" in types
        assert "integer" in types
        assert "select_one gender" in types

    def test_parses_choices(self, xlsform_path):
        _, choices, _ = parse_xlsform(xlsform_path)
        assert "gender" in choices
        assert len(choices["gender"]) == 3
        assert choices["gender"][0]["name"] == "m"
        assert choices["gender"][0]["label"] == "Male"

    def test_parses_settings(self, xlsform_path):
        _, _, settings = parse_xlsform(xlsform_path)
        assert settings["id_string"] == "test_survey_2025"
        assert settings["version"] == "1.0"
        assert settings["default_language"] == "English"

    def test_handles_label_with_language_suffix(self, xlsform_path):
        """The fixture uses label::English — verify it's picked up correctly."""
        survey_rows, _, _ = parse_xlsform(xlsform_path)
        # extract_variables should find the label::English column
        from kobo2ddi.transform import extract_variables
        _, choices, _ = parse_xlsform(xlsform_path)
        variables = extract_variables(survey_rows, choices)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["label"] == "Full name"


# -- build_workbook ----------------------------------------------------------


class TestBuildWorkbook:
    def test_has_three_sheets(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        assert wb.sheetnames == ["variables", "data", "survey_info"]

    def test_variables_sheet_header(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["variables"]
        header = [cell.value for cell in ws[1]]
        assert header == ["name", "group", "label", "type", "values", "required", "source_type"]

    def test_variables_sheet_row_count(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["variables"]
        # header + 9 data-carrying variables
        assert ws.max_row == 10

    def test_data_sheet_columns_match_variables(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws_var = wb["variables"]
        ws_data = wb["data"]
        var_names = [ws_var.cell(row=r, column=1).value for r in range(2, ws_var.max_row + 1)]
        data_header = [cell.value for cell in ws_data[1]]
        assert data_header == var_names

    def test_data_sheet_row_count(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["data"]
        # header + 2 submissions
        assert ws.max_row == 3

    def test_data_values_mapped_correctly(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["data"]
        # Row 2 = first submission (Alice)
        header = [cell.value for cell in ws[1]]
        row_values = {header[i]: ws.cell(row=2, column=i + 1).value for i in range(len(header))}
        assert row_values["full_name"] == "Alice"
        assert row_values["age"] == "30"
        assert row_values["gender"] == "f"

    def test_survey_info_content(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("My Survey", survey_rows, choices_by_list, settings, submissions)
        ws = wb["survey_info"]
        info = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(2, ws.max_row + 1)}
        assert info["title"] == "My Survey"
        assert info["id"] == "test_survey_2025"
        assert info["version"] == "1.0"
        assert info["language"] == "English"
        assert info["source"] == "kobotoolbox"
        assert info["submission_count"] == 2
        assert info["export_date"] == date.today().isoformat()

    def test_empty_submissions(self, survey_rows, choices_by_list, settings):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, [])
        ws = wb["data"]
        assert ws.max_row == 1  # header only

    def test_roundtrip_through_file(self, tmp_path, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        path = tmp_path / "out.xlsx"
        wb.save(path)
        import openpyxl
        wb2 = openpyxl.load_workbook(path)
        assert wb2.sheetnames == ["variables", "data", "survey_info"]
