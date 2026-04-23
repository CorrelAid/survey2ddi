"""Tests for kobo2ddi.transform — variable extraction, XLSForm parsing, xlsx output."""

from datetime import date

from kobo2ddi.transform import extract_variables, parse_xlsform, build_workbook


# -- extract_variables -------------------------------------------------------


class TestExtractVariables:
    def test_returns_only_data_carrying_types(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        names = [v["name"] for v in variables]
        # data-carrying fields present
        assert "full_name" in names
        assert "age" in names
        assert "gender" in names
        assert "satisfaction" in names
        assert "thanks" in names  # note vars are emitted (matches qwacback)
        # structural rows excluded
        assert "demo" not in names  # begin_group name

    def test_group_tracking(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["group"] == "demo"
        assert by_name["satisfaction"]["group"] == "feedback"

    def test_data_key_includes_group(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["_data_key"] == "demo/full_name"
        assert by_name["satisfaction"]["_data_key"] == "feedback/satisfaction"

    def test_type_mapping(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["type"] == "string"
        assert by_name["age"]["type"] == "integer"
        assert by_name["gender"]["type"] == "select_one"
        assert by_name["hobbies"]["type"] == "select_multiple"
        assert by_name["satisfaction"]["type"] == "select_one"
        assert by_name["score"]["type"] == "range"
        assert by_name["rating"]["type"] == "decimal"
        assert by_name["visit_date"]["type"] == "date"
        assert by_name["calc_field"]["type"] == "calculate"

    def test_choices_resolved(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        gender = by_name["gender"]
        assert len(gender["choices"]) == 3
        assert gender["values"] == "m=Male|f=Female|other=Other"

    def test_no_choices_for_non_select(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["choices"] == []
        assert by_name["full_name"]["values"] == ""

    def test_source_type_preserved(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["gender"]["source_type"] == "select_one gender"
        assert by_name["full_name"]["source_type"] == "text"

    def test_measure_inference(self, survey_rows, choices_by_list):
        variables = extract_variables(survey_rows, choices_by_list)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["measure"] == ""  # string → empty
        assert by_name["age"]["measure"] == "ratio"
        assert by_name["gender"]["measure"] == "nominal"
        assert by_name["hobbies"]["measure"] == "nominal"
        assert by_name["satisfaction"]["measure"] == "ordinal"  # likert
        assert by_name["score"]["measure"] == "ratio"
        assert by_name["rating"]["measure"] == "ratio"
        assert by_name["visit_date"]["measure"] == "interval"
        assert by_name["calc_field"]["measure"] == ""  # calculate → empty

    def test_likert_detected_by_list_name(self):
        rows = [{"type": "select_one likert_scale", "name": "q1", "label": "Q1", "required": "false"}]
        choices = {"likert_scale": [{"name": "1", "label": "Low"}, {"name": "2", "label": "High"}]}
        variables = extract_variables(rows, choices)
        assert variables[0]["measure"] == "ordinal"

    def test_likert_detected_by_appearance(self):
        rows = [{"type": "select_one mylist", "name": "q1", "label": "Q1", "required": "false", "appearance": "likert"}]
        choices = {"mylist": [{"name": "1", "label": "Low"}, {"name": "2", "label": "High"}]}
        variables = extract_variables(rows, choices)
        assert variables[0]["measure"] == "ordinal"

    def test_unknown_type_passed_through(self):
        rows = [{"type": "newtype", "name": "x", "label": "X", "required": "false"}]
        variables = extract_variables(rows, {})
        assert variables[0]["type"] == "newtype"

    def test_empty_survey_rows(self):
        assert extract_variables([], {}) == []

    def test_skipped_types_not_included(self):
        """Metadata/geo/media types are filtered out; note is intentionally kept."""
        rows = [
            {"type": "geopoint", "name": "loc", "label": "Location"},
            {"type": "image", "name": "photo", "label": "Photo"},
            {"type": "start", "name": "start", "label": "Start"},
        ]
        assert extract_variables(rows, {}) == []

    def test_note_type_retained(self):
        """note vars are emitted as text vars — matches qwacback's converter."""
        rows = [{"type": "note", "name": "thanks", "label": "Thank you"}]
        variables = extract_variables(rows, {})
        assert len(variables) == 1
        assert variables[0]["name"] == "thanks"
        assert variables[0]["type"] == "note"

    def test_blank_type_row_skipped(self):
        """Rows with an empty/None type are silently skipped (line 126)."""
        rows = [
            {"type": "", "name": "phantom", "label": "Ghost"},
            {"type": "text", "name": "real", "label": "Real"},
        ]
        variables = extract_variables(rows, {})
        names = [v["name"] for v in variables]
        assert "phantom" not in names
        assert "real" in names

    def test_blank_name_row_skipped(self):
        """Rows with an empty/None name are silently skipped (line 142)."""
        rows = [
            {"type": "text", "name": "", "label": "No name"},
            {"type": "text", "name": "named", "label": "Named"},
        ]
        variables = extract_variables(rows, {})
        names = [v["name"] for v in variables]
        assert "" not in names
        assert "named" in names


# -- parse_xlsform -----------------------------------------------------------


class TestParseXlsform:
    def test_parses_survey_sheet(self, xlsform_path):
        survey_rows, _, _ = parse_xlsform(xlsform_path)
        types = [r["type"] for r in survey_rows]
        assert "text" in types
        assert "integer" in types
        assert "select_one gender" in types

    def test_parses_choices(self, xlsform_path):
        _, choices, _ = parse_xlsform(xlsform_path)
        assert "gender" in choices
        assert len(choices["gender"]) == 3
        assert choices["gender"][0]["name"] == "m"
        assert choices["gender"][0]["label"] == "Male"

    def test_parses_settings(self, xlsform_path):
        _, _, settings = parse_xlsform(xlsform_path)
        assert settings["id_string"] == "test_survey_2025"
        assert settings["version"] == "1.0"
        assert settings["default_language"] == "English"

    def test_missing_choices_sheet_returns_empty(self, tmp_path):
        """XLSForm without a 'choices' sheet → choices_by_list is empty (line 59)."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "survey"
        ws.append(["type", "name", "label"])
        ws.append(["text", "q1", "Q1"])
        # no 'choices' sheet
        ws_s = wb.create_sheet("settings")
        ws_s.append(["id_string"])
        ws_s.append(["x"])
        path = tmp_path / "form.xlsx"
        wb.save(path)
        _, choices, _ = parse_xlsform(path)
        assert choices == {}

    def test_choices_row_without_list_name_skipped(self, tmp_path):
        """Choices row with blank list_name is silently skipped (line 78)."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "survey"
        ws.append(["type", "name", "label"])
        ws_c = wb.create_sheet("choices")
        ws_c.append(["list_name", "name", "label"])
        ws_c.append([None, "orphan", "Orphan choice"])   # blank list_name
        ws_c.append(["real_list", "a", "Option A"])
        wb.create_sheet("settings")
        path = tmp_path / "form.xlsx"
        wb.save(path)
        _, choices, _ = parse_xlsform(path)
        assert "real_list" in choices
        assert None not in choices
        assert "" not in choices

    def test_handles_label_with_language_suffix(self, xlsform_path):
        """The fixture uses label::English — verify it's picked up correctly."""
        survey_rows, _, _ = parse_xlsform(xlsform_path)
        # extract_variables should find the label::English column
        from kobo2ddi.transform import extract_variables
        _, choices, _ = parse_xlsform(xlsform_path)
        variables = extract_variables(survey_rows, choices)
        by_name = {v["name"]: v for v in variables}
        assert by_name["full_name"]["label"] == "Full name"


# -- build_workbook ----------------------------------------------------------


class TestBuildWorkbook:
    def test_has_three_sheets(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        assert wb.sheetnames == ["variables", "data", "survey_info"]

    def test_variables_sheet_header(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["variables"]
        header = [cell.value for cell in ws[1]]
        assert header == ["name", "group", "label", "type", "measure", "values", "required", "source_type"]

    def test_variables_sheet_row_count(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["variables"]
        # header + 10 data-carrying variables (9 + `thanks` note)
        assert ws.max_row == 11

    def test_data_sheet_columns_match_variables(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws_var = wb["variables"]
        ws_data = wb["data"]
        var_names = [ws_var.cell(row=r, column=1).value for r in range(2, ws_var.max_row + 1)]
        data_header = [cell.value for cell in ws_data[1]]
        assert data_header == var_names

    def test_data_sheet_row_count(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["data"]
        # header + 2 submissions
        assert ws.max_row == 3

    def test_data_values_mapped_correctly(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        ws = wb["data"]
        # Row 2 = first submission (Alice)
        header = [cell.value for cell in ws[1]]
        row_values = {header[i]: ws.cell(row=2, column=i + 1).value for i in range(len(header))}
        assert row_values["full_name"] == "Alice"
        assert row_values["age"] == "30"
        assert row_values["gender"] == "f"

    def test_survey_info_content(self, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("My Survey", survey_rows, choices_by_list, settings, submissions)
        ws = wb["survey_info"]
        info = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(2, ws.max_row + 1)}
        assert info["title"] == "My Survey"
        assert info["id"] == "test_survey_2025"
        assert info["version"] == "1.0"
        assert info["language"] == "English"
        assert info["source"] == "kobotoolbox"
        assert info["submission_count"] == 2
        assert info["export_date"] == date.today().isoformat()

    def test_empty_submissions(self, survey_rows, choices_by_list, settings):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, [])
        ws = wb["data"]
        assert ws.max_row == 1  # header only

    def test_roundtrip_through_file(self, tmp_path, survey_rows, choices_by_list, settings, submissions):
        wb = build_workbook("Test", survey_rows, choices_by_list, settings, submissions)
        path = tmp_path / "out.xlsx"
        wb.save(path)
        import openpyxl
        wb2 = openpyxl.load_workbook(path)
        assert wb2.sheetnames == ["variables", "data", "survey_info"]
