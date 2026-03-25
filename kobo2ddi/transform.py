"""Transform KoboToolbox (or other) survey data into a DDI-adjacent xlsx."""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook

# Types that don't carry respondent data
SKIP_TYPES = {
    "begin_group", "end_group", "begin_repeat", "end_repeat",
    "note", "start", "end", "today", "deviceid", "phonenumber",
    "username", "audit",
    # Repeat groups are skipped entirely — variables inside begin_repeat/end_repeat
    # blocks are silently excluded.  KoboToolbox stores repeat data as nested
    # arrays; supporting them requires a different data model.
    # Unsupported for now
    "geopoint", "geotrace", "geoshape",
    "image", "audio", "video", "file", "barcode",
}

# XLSForm type → standardized type
TYPE_MAP = {
    "text": "string",
    "integer": "integer",
    "decimal": "decimal",
    "date": "date",
    "time": "time",
    "datetime": "datetime",
    "calculate": "calculate",
    "range": "range",
    "acknowledge": "acknowledge",
    "hidden": "hidden",
}

# Standardized type → measurement level (empty = leave for user to fill in)
MEASURE_MAP = {
    "select_one": "nominal",
    "select_multiple": "nominal",
    "rank": "ordinal",
    "integer": "ratio",
    "decimal": "ratio",
    "range": "ratio",
    "date": "interval",
    "time": "interval",
    "datetime": "interval",
}


def parse_xlsform(path: Path) -> tuple[list[dict], dict[str, list[dict]], dict]:
    """Parse an XLSForm xlsx into (survey_rows, choices_by_list, settings).

    Handles any language suffix on label columns (e.g. ``label::Deutsch``).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def _read_sheet(name: str) -> list[dict]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = list(rows[0])
        return [dict(zip(headers, row)) for row in rows[1:]]

    survey_rows = _read_sheet("survey")
    choices_raw = _read_sheet("choices")
    settings_rows = _read_sheet("settings")
    wb.close()

    # Build choices lookup: list_name → [{name, label}, ...]
    choices_by_list: dict[str, list[dict]] = {}
    choice_label_col = _find_label_col([r.keys() for r in choices_raw[:1]])
    for row in choices_raw:
        ln = row.get("list_name")
        if not ln:
            continue
        choices_by_list.setdefault(str(ln), []).append({
            "name": str(row.get("name", "")),
            "label": str(row.get(choice_label_col, "") or ""),
        })

    settings = settings_rows[0] if settings_rows else {}

    return survey_rows, choices_by_list, settings


def _find_label_col(header_sets: list) -> str:
    """Return the best label column name (prefer ``label::*``, fall back to ``label``).

    Limitation: for multi-language forms only the *first* ``label::*`` column is
    used.  There is currently no way to select a specific language.  If this
    matters, ensure the desired language column appears first in the XLSForm.
    """
    if not header_sets:
        return "label"
    headers = list(header_sets[0])
    for h in headers:
        if h and str(h).startswith("label::"):
            return str(h)
    return "label"


def extract_variables(
    survey_rows: list[dict],
    choices_by_list: dict[str, list[dict]],
) -> list[dict]:
    """Extract a flat list of variable dicts from parsed XLSForm data.

    Each variable dict has: name, group, label, type, values (pipe-separated),
    list_name, choices (list of dicts), required, source_type, _data_key.

    Source-agnostic — usable by both xlsx and DDI XML builders.
    """
    label_col = _find_label_col(
        [r.keys() for r in survey_rows[:1]] if survey_rows else []
    )

    variables: list[dict] = []
    group_stack: list[str] = []
    # Track group metadata: name → {label, appearance}
    group_meta: dict[str, dict] = {}

    for row in survey_rows:
        raw_type = str(row.get("type") or "").strip()
        if not raw_type:
            continue

        base_type = raw_type.split()[0]

        if base_type == "begin_group":
            gname = str(row.get("name", ""))
            group_stack.append(gname)
            group_meta[gname] = {
                "label": str(row.get(label_col, "") or ""),
                "appearance": str(row.get("appearance", "") or "").lower(),
            }
            continue
        if base_type == "end_group":
            if group_stack:
                group_stack.pop()
            continue
        if base_type in SKIP_TYPES:
            continue

        name = str(row.get("name", ""))
        if not name:
            continue

        # Determine standardized type and optional list name
        list_name = None
        if base_type in ("select_one", "select_multiple", "rank"):
            parts = raw_type.split(maxsplit=1)
            std_type = base_type
            list_name = parts[1] if len(parts) > 1 else None
        else:
            std_type = TYPE_MAP.get(base_type, base_type)

        # Resolve choices
        choices = choices_by_list.get(list_name, []) if list_name else []
        values_str = "|".join(f"{c['name']}={c['label']}" for c in choices)

        group = "/".join(group_stack) if group_stack else ""
        data_key = f"{group}/{name}" if group else name

        # Infer measurement level
        measure = MEASURE_MAP.get(std_type, "")
        if std_type == "select_one":
            appearance = str(row.get("appearance", "") or "").lower()
            if "likert" in (list_name or "").lower() or "likert" in appearance:
                measure = "ordinal"

        # Resolve immediate group metadata (innermost group)
        cur_group = group_stack[-1] if group_stack else ""
        gm = group_meta.get(cur_group, {})

        variables.append({
            "name": name,
            "group": group,
            "group_label": gm.get("label", ""),
            "group_appearance": gm.get("appearance", ""),
            "label": str(row.get(label_col, "") or ""),
            "type": std_type,
            "measure": measure,
            "list_name": list_name or "",
            "choices": choices,
            "values": values_str,
            "required": str(row.get("required", "false") or "false").lower(),
            "source_type": raw_type,
            "_data_key": data_key,
        })

    return variables


def build_workbook(
    asset_name: str,
    survey_rows: list[dict],
    choices_by_list: dict[str, list[dict]],
    settings: dict,
    submissions: list[dict],
    source: str = "kobotoolbox",
) -> Workbook:
    """Build a DDI-adjacent xlsx workbook from parsed survey data.

    Source-agnostic: LimeSurvey and other adapters pass the same inputs.
    *submissions* must be keyed by ``v["_data_key"]`` (``"group/name"`` or ``"name"``).
    *source* is written to the ``survey_info`` sheet (e.g. ``"limesurvey"``).
    """
    variables = extract_variables(survey_rows, choices_by_list)

    # --- write workbook ---
    wb = Workbook()

    # Sheet 1: variables
    ws_var = wb.active
    ws_var.title = "variables"
    var_headers = ["name", "group", "label", "type", "measure", "values", "required", "source_type"]
    ws_var.append(var_headers)
    for v in variables:
        ws_var.append([v[h] for h in var_headers])

    # Sheet 2: data
    ws_data = wb.create_sheet("data")
    col_names = [v["name"] for v in variables]
    data_keys = [v["_data_key"] for v in variables]
    ws_data.append(col_names)
    for sub in submissions:
        ws_data.append([sub.get(dk, "") for dk in data_keys])

    # Sheet 3: survey_info
    ws_info = wb.create_sheet("survey_info")
    ws_info.append(["key", "value"])
    info_rows = [
        ("title", asset_name),
        ("id", settings.get("id_string", "")),
        ("version", settings.get("version", "")),
        ("language", settings.get("default_language", "")),
        ("source", source),
        ("submission_count", len(submissions)),
        ("export_date", date.today().isoformat()),
    ]
    for k, v in info_rows:
        ws_info.append([k, v])

    return wb
